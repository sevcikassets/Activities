from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app import models
from app.repository import get_or_create_category, get_or_create_project, get_or_create_ticket


def parse_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                pass
    return None


def parse_time(value) -> time | None:
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, str) and value.strip():
        for fmt in ("%H:%M", "%H.%M"):
            try:
                return datetime.strptime(value.strip(), fmt).time()
            except ValueError:
                pass
    return None


def decimal_or_none(value) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def period_to_validity(period: str) -> tuple[datetime | None, datetime | None]:
    if not period:
        return None, None
    text = str(period).strip()
    if len(text) == 4 and text.isdigit():
        year = int(text)
        return datetime(year, 1, 1), datetime(year, 12, 31, 23, 59, 59)
    if len(text) == 7 and text[4] == "-":
        year, month = [int(part) for part in text.split("-")]
        start = datetime(year, month, 1)
        if month == 12:
            end = datetime(year, 12, 31, 23, 59, 59)
        else:
            end = datetime(year, month + 1, 1) - timedelta(seconds=1)
        return start, end
    return None, None


def import_workbook(db: Session, workbook_path: Path) -> dict:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    imported_rows = 0
    skipped_rows = 0

    if "Doprava" in wb.sheetnames:
        sheet = wb["Doprava"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name = row[0]
            if name:
                existing = db.query(models.Transport).filter(models.Transport.name == str(name)).first()
                if not existing:
                    db.add(models.Transport(name=str(name)))

    if "Režijní tikety" in wb.sheetnames:
        sheet = wb["Režijní tikety"]
        for ticket_no, project_name, subject, period, *_ in sheet.iter_rows(min_row=2, values_only=True):
            if not ticket_no:
                continue
            project = get_or_create_project(db, str(project_name) if project_name else None)
            valid_from, valid_to = period_to_validity(str(period) if period else "")
            get_or_create_ticket(
                db,
                str(ticket_no),
                project,
                subject=str(subject) if subject else None,
                is_overhead=True,
                valid_from=valid_from,
                valid_to=valid_to,
                source_period=str(period) if period else None,
            )

    if "Aktivity" not in wb.sheetnames:
        raise ValueError("Workbook does not contain sheet 'Aktivity'.")

    sheet = wb["Aktivity"]
    for index, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
        category, description, spent_on, _, _, _, transport, km, started, ended, overlap, hours, ticket_no, project_name, text, redmine_time, reported = row[1:18]
        parsed_date = parse_date(spent_on)
        duration = decimal_or_none(hours)
        if not parsed_date or duration is None or not description:
            skipped_rows += 1
            continue
        existing_entry = (
            db.query(models.TimeEntry)
            .filter(models.TimeEntry.source == "excel", models.TimeEntry.source_row == index)
            .first()
        )
        if existing_entry:
            skipped_rows += 1
            continue
        project = get_or_create_project(db, str(project_name) if project_name else None)
        ticket = get_or_create_ticket(db, str(ticket_no), project) if ticket_no else None
        category_code = get_or_create_category(db, str(category) if category else None)
        transport_row = None
        if transport:
            transport_row = db.query(models.Transport).filter(models.Transport.name == str(transport)).first()
            if not transport_row:
                transport_row = models.Transport(name=str(transport))
                db.add(transport_row)
                db.flush()
        db.add(
            models.TimeEntry(
                spent_on=parsed_date,
                started_at=parse_time(started),
                ended_at=parse_time(ended),
                duration_hours=duration,
                category_code=category_code,
                description=str(description),
                ticket_id=ticket.id if ticket else None,
                project_id=project.id if project else None,
                transport_id=transport_row.id if transport_row else None,
                km=decimal_or_none(km),
                overlap_hours=decimal_or_none(overlap) or Decimal("0"),
                redmine_time=str(redmine_time) if redmine_time else None,
                reported_status=str(reported) if reported else None,
                source="excel",
                source_row=index,
                raw_text=str(text) if text else None,
            )
        )
        imported_rows += 1
        if imported_rows % 1000 == 0:
            db.commit()

    db.commit()
    return {"imported_rows": imported_rows, "skipped_rows": skipped_rows}
